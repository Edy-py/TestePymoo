# -*- coding: utf-8 -*-
"""
Este módulo contém funções para processar dados de temperatura e manipular
arquivos de simulação do EnergyPlus (.idf) para análises de conforto térmico.

"""

import os
import sys
import calendar
import subprocess
import shutil
from datetime import datetime, timedelta
from typing import List, Any, Dict, Tuple, Optional
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --- Constantes ---
ANO_SIMULACAO = 2025
TEMP_ZONAS_BIOCLIMATICAS: Dict[str, Tuple[float, float]] = {
    'Zona 1': (27.68, 20.68),
    'Zona 2': (25.58, 18.58),
    'Zona 3': (27.18, 20.18),
    'Zona 4': (27.25, 20.25),
    'Zona 5': (26.34, 19.34),
    'Zona 6': (27.66, 20.66),
    'Zona 7': (29.81, 22.81),
    'Zona 8': (29.57, 22.57)
}
NOME_VARIAVEL_TEMPERATURA = "Mean Air Temperature"


def ajustar_largura_colunas_excel(caminho_arquivo: str) -> None:
    """Ajusta a largura e alinhamento das colunas de um arquivo Excel.

    :param caminho_arquivo: O caminho para o arquivo Excel a ser modificado.
    :type caminho_arquivo: str
    """
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    wb.save(caminho_arquivo)


def csv_para_excel(caminho_csv: str, pasta_saida: str, nome_base_idf: str, zona: str) -> Optional[str]:
    """Converte um arquivo CSV do EnergyPlus para um arquivo Excel.

    :param caminho_csv: Caminho para o arquivo CSV de entrada.
    :type caminho_csv: str
    :param pasta_saida: Diretório onde o arquivo Excel será salvo.
    :type pasta_saida: str
    :param nome_base_idf: O nome base do arquivo IDF para usar no nome do arquivo de saída.
    :type nome_base_idf: str
    :param zona: A zona bioclimática para usar no nome do arquivo de saída.
    :type zona: str
    :return: O caminho para o arquivo Excel criado, ou None em caso de erro.
    :rtype: str or None
    """
    try:
        dados_csv = pd.read_csv(caminho_csv)
        nome_excel = f"Resultados_Excel_{nome_base_idf.strip(".idf")}_{zona}.xlsx"
        caminho_excel = os.path.join(pasta_saida, nome_excel)
        dados_csv.to_excel(caminho_excel, index=False)
        return caminho_excel
    except FileNotFoundError:
        print(f"Erro: Arquivo CSV não encontrado em '{caminho_csv}'")
        return None
    except Exception as e:
        print(f"Erro ao converter CSV para Excel: {e}")
        return None


def formatar_data_hora(df: pd.DataFrame, ano: int = ANO_SIMULACAO) -> pd.DataFrame:
    """Separa a coluna 'Date/Time' em 'Data' e 'Hora', formatando-as.

    :param df: DataFrame contendo a coluna 'Date/Time' do EnergyPlus.
    :type df: pd.DataFrame
    :param ano: O ano a ser usado para compor as datas. Padrão é ANO_SIMULACAO.
    :type ano: int
    :return: DataFrame modificado com as novas colunas 'Data' e 'Hora'.
    :rtype: pd.DataFrame
    """
    datas_formatadas, horas_formatadas = [], []

    for valor in df['Date/Time'].astype(str):
        partes = valor.strip().split()
        data_str, hora_str = (partes[0], partes[1]) if len(partes) == 2 else (partes[0], '00:00:00')

        partes_data = data_str.split('/')
        if len(partes_data) == 2:
            data_formatada = f"{partes_data[1]}/{partes_data[0]}"
            data_formatada = f"{data_formatada}/{ano}"
        else:
            data_formatada = f"{partes_data[1]}/{partes_data[0]}/{partes_data[2]}"
            
        datas_formatadas.append(data_formatada)
        horas_formatadas.append(hora_str)

    df = df.drop(columns=['Date/Time'])
    df.insert(0, "Hora", horas_formatadas)
    df.insert(0, "Data", datas_formatadas)
    return df


def converter_para_datetime(df: pd.DataFrame) -> List[pd.Timestamp]:
    """Converte colunas de string 'Data' e 'Hora' para objetos datetime.
    
    Esta função lida com o formato de hora "24:00:00" do EnergyPlus,
    convertendo-o para o dia seguinte às 00:00:00.

    :param df: DataFrame com as colunas 'Data' e 'Hora' em formato de string.
    :type df: pd.DataFrame
    :return: Uma lista de objetos pd.Timestamp.
    :rtype: list[pd.Timestamp]
    """
    datas_reais = []
    for _, row in df.iterrows():
        try:
            data_str, hora_str = row['Data'], row['Hora']
            dia, mes, ano = map(int, data_str.split('/'))
            
            if hora_str.startswith('24:'):
                data_base = datetime(ano, mes, dia) + timedelta(days=1)
                hora_str = '00' + hora_str[2:]
            else:
                data_base = datetime(ano, mes, dia)

            hora, minuto, segundo = map(int, hora_str.split(':'))
            data_completa = data_base.replace(hour=hora, minute=minuto, second=segundo)
            datas_reais.append(pd.Timestamp(data_completa))
        except (ValueError, TypeError):
            datas_reais.append(pd.NaT)
    return datas_reais


def processar_dados_temporais(df: pd.DataFrame, colunas_temp: List[str]) -> pd.DataFrame:
    """Prepara o DataFrame para análises temporais.

    Converte colunas de temperatura para numérico e cria uma coluna 'DateTime'
    com base nas colunas de data e hora formatadas.

    :param df: DataFrame a ser processado.
    :type df: pd.DataFrame
    :param colunas_temp: Lista dos nomes das colunas que contêm dados de temperatura.
    :type colunas_temp: list[str]
    :return: DataFrame processado e pronto para análise.
    :rtype: pd.DataFrame
    """
    for col in colunas_temp:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    
    df['DateTime'] = converter_para_datetime(df)
    df.dropna(subset=['DateTime'], inplace=True)
    df = df.copy()

    if df.empty:
        print("Nenhuma data válida encontrada após o processamento!")
        return df

    df['Month'] = df['DateTime'].dt.month
    df['Month_Name'] = df['DateTime'].dt.month_name()
    df['Hour'] = df['DateTime'].dt.hour
    return df


def calcular_metricas_conforto(
    dados_periodo: pd.DataFrame, 
    coluna_temp: str, 
    temp_min_conforto: float, 
    temp_max_conforto: float, 
    intervalo_minutos: int = 15
) -> Dict[str, float]:
    """Calcula as métricas de conforto térmico para um determinado período.

    :param dados_periodo: DataFrame filtrado para o período de análise (ex: um mês).
    :type dados_periodo: pd.DataFrame
    :param coluna_temp: Nome da coluna de temperatura a ser avaliada.
    :type coluna_temp: str
    :param temp_min_conforto: Limite inferior da temperatura de conforto.
    :type temp_min_conforto: float
    :param temp_max_conforto: Limite superior da temperatura de conforto.
    :type temp_max_conforto: float
    :param intervalo_minutos: O intervalo de tempo de cada registro em minutos. Padrão é 15.
    :type intervalo_minutos: int
    :return: Um dicionário com as métricas calculadas.
    :rtype: dict[str, float]
    """
    horas_por_intervalo = intervalo_minutos / 60
    total_horas = len(dados_periodo) * horas_por_intervalo

    if total_horas == 0:
        return {'Total de Horas': 0, 'Total Conforto': 0, '% Conforto': 0}

    em_conforto = dados_periodo[coluna_temp].between(temp_min_conforto, temp_max_conforto)
    total_conforto = em_conforto.sum() * horas_por_intervalo

    return {
        'Total de Horas': round(total_horas, 2),
        'Total Conforto': round(total_conforto, 2),
        '% Conforto': round((total_conforto / total_horas) * 100, 1) if total_horas else 0
    }


def analisar_temperaturas(df: pd.DataFrame, temp_min: float, temp_max: float) -> pd.DataFrame:
    """Orquestra a análise de temperatura e conforto do DataFrame.

    :param df: O DataFrame completo com os resultados da simulação.
    :type df: pd.DataFrame
    :param temp_min: Temperatura mínima de conforto para a análise.
    :type temp_min: float
    :param temp_max: Temperatura máxima de conforto para a análise.
    :type temp_max: float
    :return: Um novo DataFrame contendo o relatório de conforto mensal.
    :rtype: pd.DataFrame
    """
    df = formatar_data_hora(df)
    colunas_temp = [col for col in df.columns if NOME_VARIAVEL_TEMPERATURA in col]
    df = processar_dados_temporais(df, colunas_temp)

    if df.empty:
        print("Nenhum dado processado para gerar relatório.")
        return pd.DataFrame()

    resultados = []
    for mes_num in range(1, 13):
        dados_mes = df[df['Month'] == mes_num]
        for temp_col in colunas_temp:
            metricas = calcular_metricas_conforto(dados_mes, temp_col, temp_min, temp_max)
            resultados.append({
                'Mês': calendar.month_name[mes_num],
                'Local': temp_col,
                **metricas
            })
            
    return pd.DataFrame(resultados)


def executar_energyplus(caminho_ep: str , pasta_simulacao: str, caminho_epw: str, caminho_idf: str) -> str:
    """Executa a simulação do EnergyPlus e limpa arquivos desnecessários.

    :param pasta_simulacao: Diretório onde a simulação será executada e os resultados salvos.
    :param caminho_epw: Caminho para o arquivo de clima (.epw).
    :param caminho_ep: Caminho para o executável do EnergyPlus.
    :param caminho_idf: Caminho para o arquivo do modelo da edificação (.idf).
    :raises sys.exit: Encerra o script se o EnergyPlus retornar um código de erro.
    :return: O caminho para o arquivo CSV de resultados gerado pelo EnergyPlus.
    """

    
  
    comando = [
        caminho_ep, "-r", "-d", pasta_simulacao,
        "-w", caminho_epw, "--expandobjects", "--readvars", caminho_idf
    ]
    print(f"Rodando comando: {' '.join(comando)}")
    resultado = subprocess.run(comando, capture_output=True, text=True, check=False)
    
    if resultado.returncode != 0:
        print(f"Erro ao rodar EnergyPlus:\n{resultado.stderr}")
        flag =  1
    else:
        flag = 0

        # sys.exit(1)

    # Limpa a pasta, mantendo apenas o CSV de saída principal
    # for nome_arquivo in os.listdir(pasta_simulacao):
    #     if nome_arquivo != "eplusout.csv":
    #         os.remove(os.path.join(pasta_simulacao, nome_arquivo))
    
    nome_arquivo_csv = caminho_idf.replace(".idf", ".csv")
    # shutil.copyfile(os.path.join(pasta_simulacao, "eplusout.csv"), os.path.join(pasta_simulacao, nome_arquivo_csv))
    # os.remove(os.path.join(pasta_simulacao, "eplusout.csv"))

    return os.path.join(pasta_simulacao, nome_arquivo_csv), flag


def processar_idf(pasta_saida: str, zona_selecionada: str, caminho_ep: str, caminho_epw: str, caminho_idf: str) -> Dict[str, float]:
    """
    Função principal responsável por execultar todas as funçoes de processamento e análise do IDF.

    :param pasta_saida: Diretório para salvar os relatórios finais.
    :param zona_selecionada: O nome da zona bioclimática para usar na análise.
    :param caminho_ep: caminho para o executável do EnergyPlus.
    :param caminho_epw: caminho para o arquivo EPW com dados climáticos.
    :param caminho_idf: caminho para o arquivo IDF a ser analisado.
    :return: Um dicionário com o total de conforto por cômodo.
    """
    
    # Aqui podemos colocar a montagem do energyplus
    # alterar_versao_energyplus_idf(idf_base, idf_alter, "24.1")
    # configurar_simulation_control(idf_alter, idf_alter)
    # configurar_building(idf_alter, idf_alter)
    # configurar_timestep(idf_alter, idf_alter)
    # adicionar_localizacao_do_epw(idf_alter, idf_alter, epw)
    # configurar_run_period(idf_alter, idf_alter)
    # configurar_dias_especiais(idf_alter, idf_alter)
    # adicionar_materiais_do_excel(idf_alter, idf_alter, excel)
    # configurar_output_variable(idf_alter, idf_alter)
    # falta a parte de montar o contruction do EnergyPlus e a parte da janela
    
    
    temp_max, temp_min = TEMP_ZONAS_BIOCLIMATICAS[zona_selecionada]
    
    caminho_csv = executar_energyplus(caminho_ep, pasta_saida, caminho_epw, caminho_idf)

    caminho_excel_resultados = csv_para_excel(caminho_csv, pasta_saida, caminho_idf, zona_selecionada)
    if not caminho_excel_resultados:
        sys.exit(1) # finaliza o script se não houver resultados
    
    df_resultados = pd.read_excel(caminho_excel_resultados)
    os.remove(caminho_csv) # Remove o CSV temporário após a conversão para Excel

    df_conforto = analisar_temperaturas(df_resultados, temp_min, temp_max)
    
    nome_tabela = f"Tabela_conforto_{caminho_idf.strip(".idf")}_{zona_selecionada}.xlsx"

    caminho_tabela = os.path.join(pasta_saida, nome_tabela)
    df_conforto.to_excel(caminho_tabela, index=False)

    ajustar_largura_colunas_excel(caminho_tabela)
    
    # Gera o relatório de texto resumido por cômodo
    df_conforto['Comodo'] = df_conforto['Local'].str.split(':', expand=True)[0]

    total_conforto_por_comodo = df_conforto.groupby('Comodo')['Total Conforto'].sum() # Agrupa por cômodo e soma o total de conforto
    
    linhas_relatorio = [
        f"- Cômodo '{comodo}': {total:.1f} horas de conforto\n"
        for comodo, total in total_conforto_por_comodo.items()
    ]

    Dict_total_conforto_por_comodo = total_conforto_por_comodo.to_dict()


    bloco_relatorio = (
        f"--- Relatório: {caminho_idf.strip(".idf")} | Zona: {zona_selecionada} ---\n"
        f"{''.join(linhas_relatorio)}"
        f"--------------------------------------------------\n\n"
    )

    caminho_txt = os.path.join(pasta_saida, "Relatorio_conforto_geral.txt")
    with open(caminho_txt, "a", encoding="utf-8") as f:
        f.write(bloco_relatorio)

    return Dict_total_conforto_por_comodo['QUARTO']


# ==============================================================================
# 4. PARTE 2: MANIPULAÇÃO DE ARQUIVOS IDF DO ENERGYPLUS
# ==============================================================================

def substituir_secao_idf(entrada_idf: str, tipo_alteracao: str, marcador_inicio: str, marcador_fim: str, novo_conteudo: list) -> str:
    """
    Função auxiliar para substituir o conteúdo entre dois marcadores em um arquivo de entrada do Energy Plus (.idf) e salva em um novo arquivo.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)
    :param tipo_alteracao: Tipo de alteração sendo feita (usado para nomear o arquivo de saída)
    :param marcador_inicio: Texto que identifica o início da seção
    :param marcador_fim: Texto que identifica o fim da seção
    :param novo_conteudo: Lista de strings para inserir na seção

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    
    with open(entrada_idf, 'r', encoding='utf-8') as f:
        linhas = f.readlines()
    idx_inicio = next(i for i, l in enumerate(linhas) if marcador_inicio in l)
    idx_fim = next(i for i, l in enumerate(linhas) if marcador_fim in l)
    linhas_novas = linhas[:idx_inicio + 1] + novo_conteudo + linhas[idx_fim:]
    nome_base = os.path.splitext(entrada_idf)[0]
    caminho_saida = f"{nome_base}_modificado_{tipo_alteracao}.idf"
    with open(caminho_saida, 'w', encoding='utf-8') as f:
        f.writelines(linhas_novas)

    return caminho_saida


def alterar_versao_energyplus_idf(entrada_idf: str, versao_instalada: str = '24.1') -> str:
    """
    Configura o objeto 'Version' em um arquivo de entrada do Energy Plus (.idf). Salva em um novo arquivo de mesmo nome acrescido de '_modificado_versao'.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)
    :param versao_instalada: String da versão instalada (padrão '24.1').

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """

    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: VERSION ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: SIMULATIONCONTROL ==========="
    conteudo_novo = [
        '\n',
        'Version,\n',
        f'    {versao_instalada};                    !- Version Identifier\n',
        '\n\n'
    ]
    caminho_saida_versao = substituir_secao_idf(entrada_idf, "versao", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida_versao


def configurar_simulation_control(entrada_idf: str) -> str:
    """
    Configura o objeto 'SimulationControl' em um arquivo de entrada do Energy Plus (.idf). Salva em um novo arquivo de mesmo nome acrescido de '_modificado_simulation_control'.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: SIMULATIONCONTROL ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: BUILDING ==========="
    conteudo_novo = [
        '\n',
        'SimulationControl,\n',
        '    Yes,                      !- Do Zone Sizing Calculation\n',
        '    Yes,                      !- Do System Sizing Calculation\n',
        '    Yes,                      !- Do Plant Sizing Calculation\n',
        '    Yes,                      !- Run Simulation for Sizing Periods\n',
        '    Yes,                      !- Run Simulation for Weather File Run Periods\n',
        '    No,                       !- Do HVAC Sizing Simulation for Sizing Periods\n',
        '    1;                        !- Maximum Number of HVAC Sizing Simulation Passes\n',
        '\n\n'
    ]
    caminho_saida_simulation = substituir_secao_idf(entrada_idf, "simulation_control", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida_simulation


def configurar_building(entrada_idf: str) -> str:
    """
    Configura o objeto 'Building' em um arquivo de entrada do Energy Plus (.idf). Salva em um novo arquivo de mesmo nome acrescido de '_modificado_building'.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: BUILDING ==========="
    marcador_fim = "===========  ALL OBJECTS IN CLASS: TIMESTEP ==========="
    conteudo_novo = [
        '\nBuilding,\n',
        '    Novo,                     !- Name\n',
        '    -60,                      !- North Axis {deg}\n',
        '    Urban,                    !- Terrain\n',
        '    0.04,                     !- Loads Convergence Tolerance Value {W}\n',
        '    0.4,                      !- Temperature Convergence Tolerance Value {deltaC}\n',
        '    FullInteriorAndExterior,  !- Solar Distribution\n',
        '    25,                       !- Maximum Number of Warmup Days\n',
        '    6;                        !- Minimum Number of Warmup Days\n',
        '\n\n'
    ]
    caminho_saida_building = substituir_secao_idf(entrada_idf, "building", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida_building


def configurar_timestep(entrada_idf: str) -> str:
    """
    Configura o objeto 'TimeStep' em um arquivo de entrada do Energy Plus (.idf). Salva em um novo arquivo de mesmo nome acrescido de '_modificado_timestep'.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: TIMESTEP ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: SITE:LOCATION ==========="
    conteudo_novo = [
        '\nTimestep,\n',
        '    4;                        !- Number of Timesteps per Hour\n',
        '\n\n'
    ]
    caminho_saida_timestep = substituir_secao_idf(entrada_idf, "timestep", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida_timestep


def configurar_run_period(entrada_idf: str) -> str:
    """
    Configura o objeto 'RunPeriod' em um arquivo de entrada do Energy Plus (.idf). Salva em um novo arquivo de mesmo nome acrescido de '_modificado_run_period'.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: RUNPERIOD ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: RUNPERIODCONTROL:SPECIALDAYS ==========="
    conteudo_novo = [
        '\nRunPeriod,\n',
        '    Run Period 1,             !- Name\n',
        '    1,                        !- Begin Month\n',
        '    1,                        !- Begin Day of Month\n',
        '    ,                         !- Begin Year\n',
        '    12,                       !- End Month\n',
        '    31,                       !- End Day of Month\n',
        '    ,                         !- End Year\n',
        '    Wednesday,                !- Day of Week for Start Day\n',
        '    Yes,                      !- Use Weather File Holidays and Special Days\n',
        '    No,                       !- Use Weather File Daylight Saving Period\n',
        '    No,                       !- Apply Weekend Holiday Rule\n',
        '    Yes,                      !- Use Weather File Rain Indicators\n',
        '    Yes;                      !- Use Weather File Snow Indicators\n',
        '\n\n'
    ]
    caminho_saida_run = substituir_secao_idf(entrada_idf, "run_period", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida_run


def configurar_dias_especiais(entrada_idf: str) -> str:
    """
    Configura o objeto 'RunPeriodControl:SpecialDays' em um arquivo de entrada do Energy Plus (.idf). Salva em um novo arquivo de mesmo nome acrescido de '_modificado_special_days'.
    
    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: RUNPERIODCONTROL:SPECIALDAYS ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: RUNPERIODCONTROL:DAYLIGHTSAVINGTIME ==========="
    feriados = [
        ('New Years Day', 'January 1'), ('Carnaval1', 'March 4'), ('Carnaval2', 'March 5'),
        ('Paixao de Cristo', 'April 19'), ('Tiradentes', 'April 21'), ('Dia do Trabalho', 'May 1'),
        ('Corpus Christi', 'June 20'), ('Independencia', 'September 7'), ('Padroeira', 'October 12'),
        ('Finados', 'November 2'), ('Proclamacao Republica', 'November 15'), ('Natal', 'December 25')
    ]
    
    conteudo_novo = ['\n']
    for nome, data in feriados:
        conteudo_novo.extend([
            'RunPeriodControl:SpecialDays,\n',
            f'    {nome},               !- Name\n',
            f'    {data},               !- Start Date\n',
            '    1,                        !- Duration {days}\n',
            '    Holiday;                  !- Special Day Type\n',
            '\n'
        ])
    conteudo_novo.append('\n')
    caminho_saida_feriados = substituir_secao_idf(entrada_idf, "feriados", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida_feriados


def configurar_localizacao_do_epw(entrada_idf: str, entrada_epw: str) -> str:
    """
    Extrai dados de localização de um arquivo EPW e configura o objeto 'Location' em um arquivo de entrada do Energy Plus (.idf). Salva em um novo arquivo de mesmo nome acrescido de '_modificado_location'.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf)
    :param entrada_epw: Caminho do arquivo climático .epw.

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    
    with open(entrada_epw, 'r', encoding='latin-1') as f:    # EPW usa codificação latin-1 por isso eu alterei aqui
        linha_loc = next(l for l in f if l.startswith("LOCATION"))
    partes = linha_loc.strip().split(',')
    cidade = partes[1]
    latitude = partes[6]
    longitude = partes[7]
    timezone = partes[8]
    elevacao = partes[9]

    conteudo_novo = [
        '\nSite:Location,\n',
        f'    {cidade},                 !- Name\n',
        f'    {latitude},               !- Latitude {{deg}}\n',
        f'    {longitude},              !- Longitude {{deg}}\n',
        f'    {timezone},               !- Time Zone {{hr}}\n',
        f'    {elevacao};                !- Elevation {{m}}\n',
        '\n\n'
    ]
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: SITE:LOCATION ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: RUNPERIOD ==========="
    caminho_saida_epw = substituir_secao_idf(entrada_idf, "localizacao_epw", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida_epw


def configurar_output_variable(caminho_idf: str, caminho_saida: str) -> None:
    """
    Adiciona objetos 'Output:Variable' padrão a um arquivo IDF.

    :param caminho_idf: Caminho do arquivo .idf original.
    :param caminho_saida: Caminho do novo arquivo .idf modificado.
    """
    conteudo_novo = [
        '\n',
        'Output:Variable, *, Zone Mean Air Temperature, Timestep;\n',
        '\n',
        'Output:Variable, *, Site Outdoor Air Drybulb Temperature, Timestep;\n',
        '\n'
    ]
    
    with open(caminho_idf, 'r', encoding='utf-8') as f:
        linhas = f.readlines()

    marcador = "!-   ===========  ALL OBJECTS IN CLASS: OUTPUT:VARIABLE ==========="
    idx = next(i for i, l in enumerate(linhas) if marcador in l)

    linhas_novas = linhas[:idx + 1] + conteudo_novo + linhas[idx + 1:]

    with open(caminho_saida, 'w', encoding='utf-8') as f:
        f.writelines(linhas_novas)


def _formatar_bloco_material(valores: List[Any], eh_ultimo: bool) -> List[str]:
    """
    Formata um único material para o formato IDF no estilo Material,
    incluindo os comentários alinhados.
    
    :param valores: Lista de valores do material.
    :param eh_ultimo: Indica se é o último material (para evitar nova linha extra).

    :return: Lista de strings formatadas para o material.
    """

    COMENTARIOS_MATERIAL = [
    "Name",
    "Roughness",
    "Thickness {m}",
    "Conductivity {W/m-K}",
    "Density {kg/m3}",
    "Specific Heat {J/kg-K}",
    "Thermal Absorptance",
    "Solar Absorptance",
    "Visible Absorptance"
]

    bloco = ["Material,\n"]

    n = 30 

    for i, valor in enumerate(valores):
        terminador = ";" if i == len(valores) - 1 else ","
        
       
        comentario = COMENTARIOS_MATERIAL[i] if i < len(COMENTARIOS_MATERIAL) else ""

        
        if valor is None or (isinstance(valor, float) and pd.isna(valor)) or str(valor).strip() == "":
            linha_valor = f"    {terminador}"
        else:
            linha_valor = f"    {valor}{terminador}"
        
        
        linha_formatada = f"{linha_valor:<{n}}!- {comentario}\n"
        bloco.append(linha_formatada)

    if not eh_ultimo:
        bloco.append('\n')
        
    return bloco


def _ler_e_formatar_materiais_excel(caminho_excel: str) -> List[str]:
    """
    Lê materiais de um Excel e formata para o estilo IDF.
    
    :param caminho_excel: Caminho do arquivo Excel com os dados dos materiais.
    :return: Lista de strings formatadas para os materiais.
    """
    df = pd.read_excel(caminho_excel)
    
    
    lista_materiais = [
        [item for item in row] 
        for row in df.values.tolist()
    ]

    lista_formatada = ["\n"]
    total_materiais = len(lista_materiais)
    for i, material_vals in enumerate(lista_materiais):
        eh_ultimo = (i == total_materiais - 1)
        lista_formatada.extend(_formatar_bloco_material(material_vals, eh_ultimo))
    lista_formatada.append("\n\n")

    return lista_formatada


def adicionar_materiais_do_excel(caminho_idf: str, caminho_excel: str) -> str:
    """
    Adiciona materiais de um arquivo Excel a um arquivo IDF.

    :param caminho_idf: Caminho do arquivo .idf original.
    :param caminho_excel: Caminho do Excel com os dados dos materiais.

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: MATERIAL ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: MATERIAL:AIRGAP ==========="
    
    conteudo_novo = _ler_e_formatar_materiais_excel(caminho_excel)
    caminho_saida = substituir_secao_idf(caminho_idf, "material", marcador_inicio, marcador_fim, conteudo_novo)
    return caminho_saida

# def selecao_materiais_otm(df: pd.DataFrame, x: list) -> tuple[str, str, str, str, str, str, str, str, str, str]:
#     """
#     Seleciona os materiais de construção a partir de uma planilha de dados previamente informada.

#     :param df: DataFrame contendo os materiais e suas propriedades termofísicas
#     :param x: Índices para a busca dos materiais desejados na tabela. A sequência atual é: (a) Pintura externa da parede, (b) Argamassa, (c) Bloco externo da parede, (d) Pintura interna da parede, (e) Telhado, (f) Forro, (g) Piso, (h) Argamassa do piso, (i) Vidro da janela, (j) Porta

#     :return: Nomes dos materiais selecionados para a construção
#     """

#     # Sem parede interna
#     pint_ext_w_ext, arg_par, blo_w_ext, pint_int, telha, forro, piso, arg_piso, vidro, porta = x

#     # Com parede interna (aqui tem que arrumar a saída se for usar)
#     # pint_ext_w_ext, arg_par, blo_w_ext, blo_w_int, pint_int, telha, forro, piso, arg_piso, vidro, porta = x

#     # Parede externa
#     m = "Wall"
#     n = "pintura"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     pint_ext_w_ext_mat = df2.loc[pint_ext_w_ext, "Material"]
#     n = "argamassa"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     arg_par_mat = df2.loc[arg_par, "Material"]
#     n = "bloco"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     blo_w_ext_mat = df2.loc[blo_w_ext, "Material"]
#     n = "pintura"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     pint_int_mat = df2.loc[pint_int, "Material"]

#     # Tellhado
#     m = "ExteriorRoof"
#     n = "telhado"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     telha_mat = df2.loc[telha, "Material"]

#     # Forro
#     m = "InteriorCeiling"
#     n = "forro"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     forro_mat = df2.loc[forro, "Material"]

#     # Piso
#     m = "InteriorFloor"
#     n = "piso"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     piso_mat = df2.loc[piso, "Material"]

#     # Argamassa do piso
#     n = "base"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     arg_piso_mat = df2.loc[arg_piso, "Material"]

#     # Esquadrias
#     m = "ExteriorWindow"
#     n = "vidro"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     vidro_mat = df2.loc[vidro, "Material"]
#     m = "ExteriorDoor"
#     n = "porta"
#     df1 = df[df["Construção"] == m]
#     df2 = df1[df1["Sub sistema"] == n]
#     df2 = df2.reset_index(drop=True)
#     porta_mat = df2.loc[porta, "Material"]

#     return pint_ext_w_ext_mat, arg_par_mat, blo_w_ext_mat, pint_int_mat, telha_mat, forro_mat, piso_mat, arg_piso_mat, vidro_mat, porta_mat


# def idf_construction(pin_ext_w_ext: str, arg_par: str, blo_w_ext: str, pin_int: str, telha: str, forro: str, piso: str, arg_piso: str, vidro: str, porta: str) -> None:
#     """
#     Cria o modelo da construção para simulação no EnergyPlus. Aqui são definidos os materiais e camadas das paredes, telhado, forro, piso e esquadrias.

#     :param pin_ext_w_ext: Pintura externa da parede
#     :param arg_par: Argamassa
#     :param blo_w_ext: Bloco externo da parede
#     :param blo_w_int: Bloco interno da parede
#     :param pin_int: Pintura interna da parede
#     :param telha: Material do telhado
#     :param forro: Material do forro
#     :param piso: Revestimento do piso
#     :param arg_piso: Argamassa do piso
#     :param vidro: Material do vidro da janela
#     :param porta: Material da porta
    
#     :return: vazio
#     """
    
#     # Construção do arquivo de saída para Energy Plus
#     r = []
#     r.append('!-   ===========  ALL OBJECTS IN CLASS: CONSTRUCTION ===========' + '\n\n')
#     n = 40
    
#     # Parede externa
#     r.append('Construction,' + '\n')
#     r.append(f'    {"Exterior Wall,".ljust(n)}!- Nome' + '\n')
#     r.append(f'    {pin_ext_w_ext.ljust(n)}!- Pintura externa' + '\n')
#     r.append(f'    {arg_par.ljust(n)}!- Argamassa externa' + '\n')
#     r.append(f'    {blo_w_ext.ljust(n)}!- Bloco parte 1' + '\n')
#     r.append(f'    {"F04 Wall air space resistance".ljust(n)}!- Camada de ar interna' + '\n')
#     r.append(f'    {blo_w_ext.ljust(n)}!- Bloco parte 2' + '\n')
#     r.append(f'    {arg_par.ljust(n)}!- Argamassa interna' + '\n')
#     r.append(f'    {pin_int.ljust(n)}!- Pintura interna' + '\n')
#     r.append('\n')

#     # # Parede interna
#     # r.append('Construction,' + '\n')
#     # r.append(f'    {"Interior Wall,".ljust(n)}!- Nome' + '\n')
#     # r.append(f'    {pin_int.ljust(n)}!- Pintura interna' + '\n')
#     # r.append(f'    {arg_par.ljust(n)}!- Argamassa' + '\n')
#     # r.append(f'    {blo_w_int.ljust(n)}!- Bloco parte 1' + '\n')
#     # r.append(f'    {"F04 Wall air space resistance".ljust(n)}!- Camada de ar' + '\n')
#     # r.append(f'    {blo_w_int.ljust(n)}!- Bloco parte 2' + '\n')
#     # r.append(f'    {arg_par.ljust(n)}!- Argamassa' + '\n')
#     # r.append(f'    {pin_int.ljust(n)}!- Pintura interna' + '\n')
#     # r.append('\n')

#     # Telhado
#     r.append('Construction,' + '\n')
#     r.append(f'    {"Exterior Roof,".ljust(n)}!- Nome' + '\n')
#     r.append(f'    {telha.ljust(n)}!- Telhado' + '\n')
#     r.append('\n')

#     # Forro
#     r.append('Construction,' + '\n')
#     r.append(f'    {"Interior Ceiling,".ljust(n)}!- Nome' + '\n')
#     r.append(f'    {forro.ljust(n)}!- Forro' + '\n')
#     r.append('\n')

#     # Piso
#     r.append('Construction,' + '\n')
#     r.append(f'    {"Interior Floor,".ljust(n)}!- Nome' + '\n')
#     r.append(f'    {piso.ljust(n)}!- Revestimento' + '\n')
#     r.append(f'    {arg_piso.ljust(n)}!- Argamassa' + '\n')
#     r.append('\n')

#     # Esquadrias
#     r.append('Construction,' + '\n')
#     r.append(f'    {"Exterior Window,".ljust(n)}!- Nome' + '\n')
#     r.append(f'    {vidro.ljust(n)}!- Vidro' + '\n')
#     r.append('\n')

#     r.append('Construction,' + '\n')
#     r.append(f'    {"Exterior Door,".ljust(n)}!- Nome' + '\n')
#     r.append(f'    {porta.ljust(n)}!- Porta' + '\n')
#     r.append('\n')

#     with open('construcao_output.txt', 'w', encoding='utf-8') as file:
#         file.writelines(r)



def criar_bloco_construction(nome: str, camadas: list):
        bloco = [f"Construction,\n", f"    {nome},\n"]
        total_camadas = len(camadas)
        for i, camada in enumerate(camadas):
            terminador = ";" if i == total_camadas - 1 else ","
            bloco.append(f"    {camada}{terminador}  !- Layer {i+1}\n")
        bloco.append("\n")
        return bloco


def configurar_construction(entrada_idf: str, args_comb: tuple, args_x: tuple) -> str:
    """
    Configura os objetos 'Construction' em um arquivo IDF com base nos índices
    selecionados pelo otimizador. Salva em um novo arquivo.

    :param entrada_idf: Caminho para o arquivo do modelo da edificação (.idf).
    :param args_comb: Tupla contendo os DataFrames com as combinações (comb_paredes, comb_radier, comb_cobertura, etc.).
    :param args_x: Tupla contendo os índices da solução atual (indice_parede, indice_radier, etc.).

    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado.
    """
    
    # --- 1. Desempacotar os argumentos ---
    # Adicione outros dataframes e índices aqui conforme necessário
    comb_paredes, comb_radier = args_comb
    indice_parede, indice_radier = args_x
    
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: CONSTRUCTION ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: GLOBALGEOMETRYRULES ==========="
    
    # --- 2. Montar as Construções Dinamicamente ---
    
    # Pega a linha (Series) correspondente à opção de parede escolhida
    parede_escolhida = comb_paredes.iloc[indice_parede]

    # Lista de camadas da parede externa (em ordem de fora para dentro)
    # pd.isna() é mais seguro pois funciona para NaN, None, etc.
    camadas_parede = [
        parede_escolhida.get('pintura externa'),
        parede_escolhida.get('revestimento externo'),
        parede_escolhida.get('bloco'),
        # Adicione aqui camadas opcionais se existirem no seu CSV
        # ex: parede_escolhida.get('camada de ar'),
        parede_escolhida.get('revestimento interno'),
        parede_escolhida.get('pintura interna')
    ]
    
    # Filtra apenas as camadas que realmente existem (ignora valores vazios/NaN)
    camadas_parede_validas = [camada for camada in camadas_parede if pd.notna(camada) and str(camada).strip()]
    
    # Pega os nomes dos materiais escolhidos
    radier_escolhido = comb_radier['Nome'].iloc[indice_radier]


    # --- 3. Gerar o Conteúdo para o IDF ---

    conteudo_novo = ["\n"] # Inicia a lista que será escrita no arquivo

    # Função auxiliar para adicionar um bloco de construção à lista
    

    # Adiciona as construções à lista 'conteudo_novo'
    conteudo_novo.extend(criar_bloco_construction("Exterior Wall", camadas_parede_validas))
    
    # Construções baseadas em Radier
    conteudo_novo.extend(criar_bloco_construction("Exterior Floor", [radier_escolhido, 'ArgamassaCimento']))
    conteudo_novo.extend(criar_bloco_construction("Interior Floor", [radier_escolhido, 'ArgamassaCimento', 'Piso Ceramico']))

    # Construções de Cobertura (lendo as camadas do CSV)
    
    
    
    # Construções Fixas (se não forem parte da otimização)
    conteudo_novo.extend(criar_bloco_construction("Interior Wall", ["ArgamassaCimento", "BlocoCeramico9cm", "ArgamassaCimento"]))
    conteudo_novo.extend(criar_bloco_construction("Interior Ceiling", ['ConcretoMacico10cm']))
    
   
    
    conteudo_novo.extend(criar_bloco_construction("Interior Window", ['VidroSimplesincolor3mm']))
    conteudo_novo.extend(criar_bloco_construction("Exterior Door", ['Porta madeira']))
    conteudo_novo.extend(criar_bloco_construction("Interior Door", ['Porta madeira']))


    # print(f"conteudo novo: {conteudo_novo}")

    # with open("conteudo_novo.txt", 'w', encoding='utf-8') as f:
    #     f.writelines(conteudo_novo)
    # --- 4. Salvar o Novo Arquivo IDF ---
    caminho_saida = substituir_secao_idf(entrada_idf, "construction", marcador_inicio, marcador_fim, conteudo_novo)

    return caminho_saida


def _formatar_bloco_vidros(valores: List[Any], eh_ultimo: bool) -> List[str]:
    """
    Formata um único material para o formato IDF no estilo WindowMaterial:Glazing,
    incluindo os comentários alinhados.
    
    :param valores: Lista de valores do material.
    :param eh_ultimo: Indica se é o último material (para evitar nova linha extra).

    :return: Lista de strings formatadas para o material.
    """

    COMENTARIOS_GLAZING = [
    "Name",
    "Optical Data Type",
    "Window Glass Spectral Data Set Name",
    "Thickness {m}",
    "Solar Transmittance at Normal Incidence",
    "Front Side Solar Reflectance at Normal Incidence",
    "Back Side Solar Reflectance at Normal Incidence",
    "Visible Transmittance at Normal Incidence",
    "Front Side Visible Reflectance at Normal Incidence",
    "Back Side Visible Reflectance at Normal Incidence",
    "Infrared Transmittance at Normal Incidence",
    "Front Side Infrared Hemispherical Emissivity",
    "Back Side Infrared Hemispherical Emissivity",
    "Conductivity {W/m-K}",
]
    
    
    bloco = ["WindowMaterial:Glazing,\n"]
    
    n = 25

    for i, valor in enumerate(valores):
        terminador = ";" if i == len(valores) - 1 else ","
        
        # Pega o comentário correspondente. Se não houver, deixa em branco.
        comentario = COMENTARIOS_GLAZING[i] if i < len(COMENTARIOS_GLAZING) else ""

        # Monta a parte do valor da linha
        if valor is None or (isinstance(valor, float) and pd.isna(valor)) or str(valor).strip() == "":
            linha_valor = f"    {terminador}"
        else:
            linha_valor = f"    {valor}{terminador}"
        
        linha_formatada = f"{linha_valor.ljust(n)}!- {comentario}\n"
        bloco.append(linha_formatada)

    
    bloco.append("\n\n")
        
    return bloco


def _ler_e_formatar_vidros_excel(caminho_excel: str) -> List[str]:
    """
    Lê materiais de um Excel e formata para o estilo IDF (WindowMaterial:Glazing etc).
    
    :param caminho_excel: Caminho do arquivo Excel com os dados dos materiais.

    :return: Lista de strings formatadas para os materiais.
    """
    df = pd.read_excel(caminho_excel)
    
    lista_materiais = [
        [item for item in row]
        for row in df.values.tolist()
    ]

    lista_formatada = ["\n"]
    total_materiais = len(lista_materiais)
    for i, material_vals in enumerate(lista_materiais):
        eh_ultimo = (i == total_materiais - 1)
        lista_formatada.extend(_formatar_bloco_vidros(material_vals, eh_ultimo))
    lista_formatada.append("\n\n")
    
    return lista_formatada


def adicionar_vidros_do_excel(caminho_idf: str, caminho_excel: str) -> str:
    """
    Adiciona vidros de um arquivo Excel a um arquivo IDF.

    :param caminho_idf: Caminho do arquivo .idf original.
    :param caminho_excel: Caminho do Excel com os dados dos materiais.
    
    :return: Caminho para o arquivo do modelo da edificação (.idf) alterado
    """
    marcador_inicio = "!-   ===========  ALL OBJECTS IN CLASS: WINDOWMATERIAL:GLAZING ==========="
    marcador_fim = "!-   ===========  ALL OBJECTS IN CLASS: WINDOWMATERIAL:GAS ==========="
    
    conteudo_novo = _ler_e_formatar_vidros_excel(caminho_excel)
    caminho_saida = substituir_secao_idf(caminho_idf, "glass", marcador_inicio, marcador_fim, conteudo_novo)
    return caminho_saida